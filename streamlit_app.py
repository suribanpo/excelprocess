import streamlit as st
import pandas as pd
import unicodedata
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule
import traceback
import sys
import datetime
import pytz

st.set_page_config(
    page_title="엑셀 데이터 통합 및 처리",
    page_icon="📑",
    layout="wide"
)

def normalize_text(value):
    if isinstance(value, str):
        return unicodedata.normalize('NFC', value)
    return value

def extract_fields(sheet_name):
    parts = sheet_name.split('_', 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    else:
        return parts[0], ""

def process_uploaded_files(uploaded_files):
    processed_files_data = {}
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        for uploaded_file in uploaded_files:
            file_name = uploaded_file.name
            try:
                base_sheet_name = '_'.join(file_name.split('_')[:2])
                excel_file = pd.ExcelFile(uploaded_file)
                sheet_dfs = []
                for sheet_name in excel_file.sheet_names:
                    try:
                        df = excel_file.parse(sheet_name=sheet_name, header=None)
                        name_row_index = df[df.apply(lambda row: row.astype(str).str.contains('이름|성명').any(), axis=1)].index[0]
                        df.columns = df.iloc[name_row_index].str.replace('성명', '이름')
                        df = df[name_row_index + 1:]
                        if '학년' in df.columns:
                            df['학년'] = df['학년'].astype(str).str.extract('(\d+)').astype(int)
                        df['영역'] = base_sheet_name
                        if len(excel_file.sheet_names) == 1:
                            new_sheet_name = base_sheet_name[:31]
                        else:
                            new_sheet_name = f"{base_sheet_name}_{sheet_name}"[:31]
                        df.to_excel(writer, sheet_name=new_sheet_name, index=False)
                        sheet_dfs.append((new_sheet_name, df))
                    except Exception as e:
                        exc_type, exc_value, exc_traceback = sys.exc_info()
                        tb_lines = traceback.format_exception(exc_type, exc_value, exc_traceback)
                        st.error(f"에러 발생! 파일: {file_name}, 시트: {sheet_name}\n{''.join(tb_lines)}")
                        return None, None
                processed_files_data[file_name] = sheet_dfs
            except Exception as e:
                exc_type, exc_value, exc_traceback = sys.exc_info()
                tb_lines = traceback.format_exception(exc_type, exc_value, exc_traceback)
                st.error(f"에러 발생! 파일: {file_name}\n{''.join(tb_lines)}")
                return None, None
    output.seek(0)
    return output, processed_files_data

def process_step2_data(step1_data):
    try:
        with pd.ExcelFile(step1_data) as excel_file:
            all_data = []
            for sheet_name in excel_file.sheet_names:
                try:
                    df = excel_file.parse(sheet_name=sheet_name)
                    max_length_col = df.apply(lambda col: col.astype(str).str.len().max(), axis=0).idxmax()
                    df.columns = df.columns.str.replace(max_length_col, '기재내용', regex=False)
                    if '학년' in df.columns:
                        df['학년'] = df['학년'].astype(str).str.extract('(\d+)').astype(int)
                        df['반'] = df['반'].astype(str).str.extract('(\d+)').astype(int)
                        df['번호'] = df['번호'].astype(str).str.extract('(\d+)').astype(int)
                    if '학번' in df.columns:
                        df['학년'] = df['학번'].astype(str).str[0].astype(int)
                        df['반'] = df['학번'].astype(str).str[1:3].astype(int)
                        df['번호'] = df['학번'].astype(str).str[3:].astype(int)
                    df = df[['학년', '반', '번호', '이름', '영역', '기재내용']]
                    df['기재내용'] = df['기재내용'].apply(lambda x: x[:x.rfind('.')+1] + ' ' if isinstance(x, str) and '.' in x else x)
                    all_data.append(df)
                except Exception as e:
                    exc_type, exc_value, exc_traceback = sys.exc_info()
                    tb_lines = traceback.format_exception(exc_type, exc_value, exc_traceback)
                    st.error(f"에러 발생! 시트: {sheet_name}\n{''.join(tb_lines)}")
                    return None
            final_df = pd.concat(all_data, ignore_index=True)
            for col in ['이름', '기재내용', '영역']:
                final_df[col] = final_df[col].apply(normalize_text)
            final_df[['영역명', '세부영역명']] = final_df['영역'].apply(lambda x: pd.Series(extract_fields(x)))
            for col in ['영역명', '세부영역명']:
                final_df[col] = final_df[col].apply(normalize_text)
            final_df = final_df[['학년', '반', '번호', '이름', '영역명', '세부영역명', '기재내용']]
            return final_df
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        tb_lines = traceback.format_exception(exc_type, exc_value, exc_traceback)
        st.error(f"2단계 처리 중 에러 발생!\n{''.join(tb_lines)}")
        return None

def create_pivot_tables(final_df):
    try:
        section_df_list = []
        for section_name in final_df['영역명'].unique():
            section_df = final_df[final_df['영역명'] == section_name]
            
            # 특기사항을 그룹화 및 피벗화
            section_df = section_df.groupby(['학년', '반', '번호', '이름', '세부영역명'], as_index=False).agg({
                '기재내용': lambda x: ' | '.join(x.dropna().astype(str))
            })
            section_df_pivot = section_df.pivot(index=['학년', '반', '번호', '이름'], columns='세부영역명', values='기재내용')
            section_df_pivot.reset_index(inplace=True)

            # 명렬표와 병합하여 누락된 학생 추가
            if 'roster_df' in st.session_state and st.session_state.roster_df is not None:
                roster_df = st.session_state.roster_df.copy()
                section_df_pivot = pd.merge(
                    roster_df,
                    section_df_pivot,
                    on=['학년', '반', '번호', '이름'],
                    how='left'
                )
            
            # NaN 값을 빈 문자열로 대체
            section_df_pivot = section_df_pivot.fillna("")

            # 결과 추가
            section_df_list.append((section_name, section_df_pivot))

        return section_df_list
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        tb_lines = traceback.format_exception(exc_type, exc_value, exc_traceback)
        st.error(f"3단계 피벗 테이블 생성 중 에러 발생!\n{''.join(tb_lines)}")
        return []
    
from openpyxl.styles import Border, Side

def add_excel_formulas(section_name, df):
    try:
        output_step4 = BytesIO()
        with pd.ExcelWriter(output_step4, engine="openpyxl") as writer:
            grouped = df.groupby(['학년', '반'])
            for (grade, class_num), group_df in grouped:
                sheet_name = f"{grade}학년_{class_num}반"[:31]
                group_df = group_df.applymap(lambda x: "" if str(x).strip() == "X" else x)
                group_df.to_excel(writer, index=False, sheet_name=sheet_name)

                wb = writer.book
                ws = wb[sheet_name]

                # 열 고정 및 시작 컬럼 설정
                ws.freeze_panes = "E2"
                start_col = 5
                num_cols = len(group_df.columns) - start_col + 1
                additional_col = start_col + num_cols
                combine_col_index = additional_col + 1
                byte_col_index = combine_col_index + 1

                # 특기사항 합본 및 바이트 계산 수식 추가
                # '이름' 열에서 마지막 유효 행 찾기
                name_col_letter = get_column_letter(group_df.columns.get_loc("이름") + 1)  # '이름' 열의 열 문자
                last_name_row = max(
                    row.row for row in ws[name_col_letter] if row.value  # '이름' 열의 유효 값이 있는 행을 찾음
                )

                # 특기사항 합본 및 바이트 계산 수식 추가 (마지막 유효 행까지만 적용)
                for idx in range(2, last_name_row + 1):  # 범위를 '이름' 열의 마지막 유효 행으로 제한
                    concat_formula = "=" + "CONCATENATE(" + ",".join(
                        [f"{get_column_letter(start_col + num_cols)}{idx}"] +  # 마지막 열 먼저 추가
                        [f"{get_column_letter(col)}{idx}" for col in range(start_col, start_col + num_cols)]  # 나머지 열 추가
                    ) + ")"
                    ws[f"{get_column_letter(combine_col_index)}{idx}"] = concat_formula
                    ws[f"{get_column_letter(byte_col_index)}{idx}"] = (
                        f'=LENB({get_column_letter(combine_col_index)}{idx})*2-LEN({get_column_letter(combine_col_index)}{idx})'
                    )

                if section_name == "자율활동":
                    ws[f"{get_column_letter(additional_col)}1"] = "비고(학급임원파일과 학급활동 등은 수기로 추가해주세요. 마지막 온점 뒤 띄어쓰기 필수!)"
                elif section_name == "진로활동":
                    ws[f"{get_column_letter(additional_col)}1"] = "비고(수기로 추가할 내용을 작성해주세요. 마지막 온점 뒤 띄어쓰기 필수!)"

                # 열 이름 설정
                # ws[f"{get_column_letter(additional_col)}1"] = "비고(학급임원파일과 학급활동 등은 수기로 추가해주세요. 마지막 온점 뒤 띄어쓰기 필수!)"
                ws[f"{get_column_letter(combine_col_index)}1"] = "특기사항 합본"
                ws[f"{get_column_letter(byte_col_index)}1"] = "바이트 계산"

                # 셀 스타일 및 포맷 적용
                yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                bold_font = Font(size=14, bold=True)
                large_bold_font = Font(size=20, bold=True)  # 20포인트 굵은 글씨 추가
                center_alignment = Alignment(horizontal="center", vertical="center")

                # 마지막 세 열(비고, 특기사항 합본, 바이트 계산) 헤더에 노란색 배경과 굵은 글씨 적용
                remarks_col_letter = get_column_letter(additional_col)
                combine_col_letter = get_column_letter(combine_col_index)
                byte_col_letter = get_column_letter(byte_col_index)

                for col_letter in [remarks_col_letter, combine_col_letter, byte_col_letter]:
                    ws[f"{col_letter}1"].fill = yellow_fill
                    ws[f"{col_letter}1"].font = bold_font
                    ws[f"{col_letter}1"].alignment = Alignment(wrap_text=True)  # 텍스트 래핑 적용


                # 바이트 계산 열 서식 (모든 행에 대해 적용)
                # 모든 행에 대해 가운데 정렬, 굵게, 글씨 크기 20포인트 적용
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws[f"{byte_col_letter}{row_idx}"]
                    cell.alignment = center_alignment
                    cell.font = large_bold_font  # 20포인트 굵은 글씨
                    # 조건부 서식은 별도로 적용


                for col_letter in [remarks_col_letter, combine_col_letter, byte_col_letter]:
                    ws[f"{col_letter}1"].fill = yellow_fill
                    ws[f"{col_letter}1"].font = bold_font
                    ws[f"{col_letter}1"].alignment = Alignment(wrap_text=True)  # 텍스트 래핑 적용

                if section_name == "자율활동":
                    # 자율활동: 바이트 계산이 1500 초과 시 빨간색 조건부 서식 적용
                    ws.conditional_formatting.add(
                        f"{byte_col_letter}2:{byte_col_letter}{ws.max_row}",
                        CellIsRule(operator="greaterThan", formula=["1500"], stopIfTrue=True, fill=red_fill)
                    )
                elif section_name == "진로활동":
                    # 진로활동: 바이트 계산이 2100 초과 시 빨간색 조건부 서식 적용
                    ws.conditional_formatting.add(
                        f"{byte_col_letter}2:{byte_col_letter}{ws.max_row}",
                        CellIsRule(operator="greaterThan", formula=["2100"], stopIfTrue=True, fill=red_fill)
                    )

                # 테두리 스타일 정의
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                # 셀 높이를 2행부터 최대값으로 고정
                # 행 높이 설정
                # 행 높이 강제 설정
                for row_idx in range(2, ws.max_row + 1):
                    ws.row_dimensions[row_idx].height = 300  # 최대값

                # 열 너비 및 텍스트 정렬 설정
                for col_idx in range(start_col, byte_col_index + 1):
                    col_letter = get_column_letter(col_idx)
                    if col_idx == combine_col_index:
                        ws.column_dimensions[col_letter].width = 150
                    elif col_idx == byte_col_index:
                        ws.column_dimensions[col_letter].width = 20
                    else:
                        ws.column_dimensions[col_letter].width = 50

                    # 셀 텍스트 줄바꿈 및 상단 정렬 + 테두리 적용
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws[f"{col_letter}{row_idx}"]
                        if cell.value:  # 값이 있는 셀에만 테두리 적용
                            cell.alignment = Alignment(wrap_text=True, vertical="center")  # 텍스트 줄바꿈
                            cell.border = thin_border

                # # 행 높이 자동 조절: 바이트 수에 따라 선형적으로 증가
                # # 변환 공식: y 포인트 ≈ 0.0805 * 바이트 + 65.15
                # for row_idx in range(2, ws.max_row + 1):
                #     byte_value = ws[f"{byte_col_letter}{row_idx}"].value
                #     if byte_value is not None and isinstance(byte_value, (int, float)):
                #         # 선형 변환 공식 적용
                #         row_height = 0.2*byte_value + 100
                #         # 행 높이를 포인트 단위로 설정 (엑셀의 단위)
                #         ws.row_dimensions[row_idx].height = row_height
                #     else:
                #         # 기본 행 높이 설정
                #         ws.row_dimensions[row_idx].height = 15

        output_step4.seek(0)
        preview_data = pd.DataFrame(df.values)
        preview_data.columns = df.columns
        return output_step4, preview_data
    except Exception as e:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        tb_lines = traceback.format_exception(exc_type, exc_value, exc_traceback)
        st.error(f"4단계 수식 추가 처리 중 에러 발생! 영역명: {section_name}\n{''.join(tb_lines)}")
        return None, None



st.title("📑 엑셀 데이터 처리 앱")

# 안내 메시지
st.info("🔑 업로드하는 통합문서에는 시트가 하나만 있어야합니다. 통합문서를 시트별로 나눠서 처리하는 기능은 아직 준비중이에요!\n\n1️⃣ 파일 업로드 및 통합\t2️⃣ 데이터 처리 및 변환\t3️⃣ 피벗 테이블 생성\t4️⃣ 수식 추가 및 반별 시트 생성")

with st.expander("📋 앱 사용 설명서 - 클릭해서 펼치기", expanded=False):
    st.markdown("""
    <div style="font-size: 14px; background-color: #f9f9f9; padding: 10px; border-radius: 5px;">
    <b>📊 엑셀 데이터 처리 앱</b><br>
    학번이나 학년/반/번호/이름이 포함된 학생 데이터를 통합, 정렬, 학생별 모아보기 등의 기능을 제공합니다.<br><br>
    <b>🔑 주요 기능</b><br>
    1️⃣ 파일 업로드 및 통합<br>
    2️⃣ 데이터 처리 및 변환<br>
    3️⃣ 피벗 테이블 생성<br>
    4️⃣ 수식 추가 및 반별 시트 생성<br><br>
    <b>📂 파일명 규칙</b><br>
    - 파일명은 <code>영역명_세부파일명_기존파일명.xlsx</code> 형태를 따라야 하며 '_'(언더바)가 정확히 두 번 들어가야 합니다.<br><br>
    <b>🔒 보안 안내</b><br>
    - 업로드된 데이터는 로컬 세션 내에서만 처리되며, 외부 서버로 전송되지 않습니다.<br><br>
    <b>🔍 미리보기 및 다운로드</b><br>
    - 각 단계에서 처리 결과를 미리 확인하고 필요한 데이터를 다운로드할 수 있습니다.<br><br>
    👤 **Creator**: Subhin Hwang
    </div>
    """, unsafe_allow_html=True)

if 'step1_data' not in st.session_state:
    st.session_state.step1_data = None  
if 'step2_data' not in st.session_state:
    st.session_state.step2_data = None  
if 'step3_data' not in st.session_state:
    st.session_state.step3_data = []  
if 'step4_data' not in st.session_state:
    st.session_state.step4_data = []  
if "uploaded_files" not in st.session_state:
    st.session_state.uploaded_files = []
if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0
if 'roster_df' not in st.session_state:
    st.session_state.roster_df = None

st.subheader("1️⃣ 학생 명렬표 업로드")
col_1_1, col_1_2 = st.columns(2)
with col_1_1:
    roster_file = st.file_uploader("학생 명렬표 업로드 (학년, 반, 번호, 이름 포함)", type=["xls", "xlsx"], key="roster")
if roster_file is not None:
    try:
        roster_df = pd.read_excel(roster_file)
        # 성공 메시지: 총 학생 수 표시
        with col_1_1:
            st.success(f"✨ 총 {len(roster_df)}명 학생이 불러와졌습니다! ")
        with col_1_2:
            st.markdown(" ")
            st.write(roster_df.head(3))

        # 열 이름 '성명'을 '이름'으로 수정
        if '성명' in roster_df.columns:
            roster_df.rename(columns={'성명': '이름'}, inplace=True)
            st.success("✅ '성명' 열 이름이 '이름'으로 수정되었습니다.")

        # 학번이 있는 경우 학년, 반, 번호로 분리
        if '학번' in roster_df.columns:
            st.write("yes")
            if '학년' in roster_df.columns:

                roster_df = roster_df.drop(columns=['학년'])
            roster_df['학년'] = roster_df['학번'].astype(str).str[0].astype(int)
            roster_df['반'] = roster_df['학번'].astype(str).str[1:3].astype(int)
            roster_df['번호'] = roster_df['학번'].astype(str).str[3:].astype(int)
            with col_1_2:
                st.success("✅ '학번'이 '학년', '반', '번호'로 분리되었습니다.")

        # 열 이름이 최종적으로 학년, 반, 번호, 이름인지 확인
        required_columns = {'학년', '반', '번호', '이름'}
        if not required_columns.issubset(roster_df.columns):
            missing_columns = required_columns - set(roster_df.columns)
            st.error(f"❌ 열 이름이 올바르지 않습니다! 다음 열이 필요합니다: {', '.join(missing_columns)}")
            st.stop()  # 오류 발생 시 실행 중지

        # 열 순서 정렬
        roster_df = roster_df[['학년', '반', '번호', '이름']]
        st.session_state.roster_df = roster_df

        # 최종 미리보기 출력
        with st.expander("📋 전처리된 학생 명단 확인"):
            st.dataframe(roster_df.head(5))

    except Exception as e:
        st.error("❌ 파일 처리 중 오류가 발생했습니다. 올바른 엑셀 파일인지 확인해주세요.")
        exc_type, exc_value, exc_traceback = sys.exc_info()
        tb_lines = traceback.format_exception(exc_type, exc_value, exc_traceback)
        st.error(f"오류 메시지:\n{''.join(tb_lines)}")

else:
    st.session_state.roster_df = None

st.subheader("2️⃣ 특기사항 파일들 업로드")

uploaded_files = st.file_uploader("특기사항 엑셀 파일 업로드 (여러개 가능)", type=["xls","xlsx"], accept_multiple_files=True, key=f"file_uploader_{st.session_state.uploader_key}")
if uploaded_files:
    st.session_state.uploaded_files = uploaded_files
    output, processed_files_data = process_uploaded_files(uploaded_files)
    if output and processed_files_data:
        st.session_state.step1_data = output
        st.session_state.processed_files_data = processed_files_data
        st.success("👏 파일 업로드 및 통합 완료")
    else:
        st.error("파일 처리 오류 발생")

    # 업로드한 모든 파일을 tabs로 보기
    tab_names = [f"▸{name.split('_')[1]}" for name in processed_files_data.keys()]
    tabs = st.tabs(tab_names) 
    for i, (file_name, sheet_dfs) in enumerate(processed_files_data.items()):
        with tabs[i]:
            # st.write(f"**{file_name} 처리 결과**")
            for sheet_name, df in sheet_dfs:
                n, m = df.shape
                st.info(f"파일명 : {file_name}....총 **{n}명** ")
                st.dataframe(df, height=200)

st.subheader("3️⃣ 엑셀파일 처리하기")


# 2단계 레이아웃: 왼쪽(설명 및 업로드), 오른쪽(결과 미리보기)
step2_l, step2_r = st.columns([0.4, 0.6])

with step2_l:
    st.write("##### 2단계: 하나의 시트로 만들기")

if st.session_state.step1_data:
    # 데이터 처리 시작
    final_df = process_step2_data(st.session_state.step1_data)
    if final_df is not None:
        st.session_state.step2_data = final_df

        # 처리 결과를 표시
        with step2_r:
            st.write("**📋 처리 결과 (미리보기)**")
            st.dataframe(final_df, height=200)

        # 결과 다운로드 버튼
        output_step2 = BytesIO()
        final_df.to_excel(output_step2, index=False, engine='xlsxwriter')
        output_step2.seek(0)
        with step2_l:
            st.success("✅ **2단계 처리 완료!**")

            # 기존 다운로드 버튼: 개별 시트로 나뉜 통합 문서
            st.download_button(
                label="📥 2단계 결과 다운로드 (개별 시트 버전)",
                data=output_step2,
                file_name="통합.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # 새로운 다운로드 버튼: 모든 데이터를 하나의 시트에 통합한 버전
            output_single_sheet = BytesIO()
            final_df.to_excel(output_single_sheet, index=False, sheet_name="모든 데이터")
            output_single_sheet.seek(0)

            st.download_button(
                label="📥 2단계 결과 다운로드 (단일 시트 버전)",
                data=output_single_sheet,
                file_name="통합_단일시트.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.error("🚨 **2단계 처리 중 오류가 발생했습니다. 입력 데이터를 확인해주세요.**")
else:
    with step2_l:
        st.warning("⚠️ **1단계 결과가 없습니다. 먼저 파일을 업로드하고 통합해주세요.**")

# 3단계 레이아웃: 왼쪽(설명 및 상태), 오른쪽(결과 미리보기)
step3_l, step3_r = st.columns([0.4, 0.6])

with step3_l:
    st.write("##### 3단계: 영역별 피벗 테이블 생성")

if st.session_state.step2_data is not None:
    section_df_list = create_pivot_tables(st.session_state.step2_data)
    if section_df_list:
        st.session_state.step3_data = section_df_list

        # 처리 결과 표시
        for section_name, df in section_df_list:
            with step3_r:
                st.write("**📋 처리 결과 (미리보기)**")
                st.dataframe(df.head(10), height=200)

                # 결과 다운로드
                output_step3 = BytesIO()
                with pd.ExcelWriter(output_step3, engine="xlsxwriter") as writer:
                    df.to_excel(writer, index=False, sheet_name="특기사항")
                output_step3.seek(0)
            with step3_l:
                st.success("✅ **3단계 처리 완료!**")

                st.download_button(
                    label=f"📥 {section_name} 3단계 결과 다운로드",
                    data=output_step3,
                    file_name=f"{section_name}_피벗.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        with step3_l:
            st.error("🚨 **3단계 처리 중 오류가 발생했습니다. 입력 데이터를 확인해주세요.**")
else:
    with step3_l:
        st.warning("⚠️ **2단계 결과가 없습니다. 먼저 데이터를 처리해주세요.**")

# 4단계 레이아웃: 왼쪽(설명 및 상태), 오른쪽(결과 미리보기)
step4_l, step4_r = st.columns([0.4, 0.6])

with step4_l:
    st.write("##### 4단계: 최종본 생성 및 서식 추가")

if st.session_state.roster_df is not None and st.session_state.step3_data:
    updated_section_df_list = []
    for section_name, df in st.session_state.step3_data:
        temp_output, preview_data = add_excel_formulas(section_name, df)
        if temp_output and preview_data is not None:
            updated_section_df_list.append((section_name, preview_data))

            with step4_r:
                st.write("**📋 처리 결과 (미리보기)**")
                st.dataframe(preview_data.head(10), height=200)

                # 최종 결과 다운로드
                temp_output.seek(0)
                kst = pytz.timezone('Asia/Seoul')
                current_datetime_kst = datetime.datetime.now(kst).strftime("%Y%m%d_%H%M")
            with step4_l:
                st.success("✅ **4단계 처리 완료! 최종본이 생성되었습니다.**")

                st.download_button(
                    label=f"📥 {section_name} 최종본 다운로드",
                    data=temp_output,
                    file_name=f"{section_name}_최종본_{current_datetime_kst}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    st.session_state.step4_data = updated_section_df_list
else:
    with step4_l:
        st.warning("⚠️ **3단계 결과 또는 학생 명렬표가 없습니다. 데이터를 확인해주세요.**")


st.markdown("---")
st.markdown("""
<div style="text-align: center; margin-top: 20px;">
    <p style="font-size: 14px; color: gray;">© 2024 <strong>Excel Process</strong> made by <strong>Subhin Hwang</strong></p>
    <p style="font-size: 12px; color: gray;">Designed with ❤️ to simplify Excel workflows.</p>
    <p style="font-size: 14px; color: gray;">📧 Contact: <a href="mailto:sbhath17@gmail.com">sbhath17@gmail.com</a></p>
</div>
""", unsafe_allow_html=True)